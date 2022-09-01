
import os
import unittest
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import src.xlsx_copycull as xlsx_copycull
    from src.xlsx_copycull import find_ranges
except ImportError:
    import sys
    sys.path.append('../src/xlsx_copycull')
    import xlsx_copycull
    from xlsx_copycull import find_ranges


class FileHandler:
    """Helper class for unittest.TestCase."""
    master = Path(r"test_data\test_data.xlsx")
    temp_dir = Path(r".\test_temp")
    sheet_name = 'Sheet1'
    temp_fn = 'test_temp.xlsx'
    temp_fp = temp_dir / temp_fn

    def __init__(self):
        self.temp_wbwrapper = None

    def new_copy(self):
        """
        (Setup method.)
        Get a fresh copy.
        :return: The new WorkbookWrapper.
        """
        self.clean_up()
        self.temp_dir.mkdir(exist_ok=True)
        self.temp_wbwrapper = xlsx_copycull.WorkbookWrapper(
            wb_fp=self.master,
            output_filename=self.temp_fn,
            copy_to_dir=self.temp_dir
        )
        return self.temp_wbwrapper

    def reload_wswrapper(self, **kwargs):
        """
        (Setup method.)
        Without reloading the WorkbookWrapper, get rid of the existing
        WorksheetWrapper (if any). Reload a new WorksheetWrapper.
        :param kwargs: Optional kwargs for `.stage_ws()`
        :return: The new WorksheetWrapper.
        """
        # Pop / discard existing keys and get a new WorksheetWrapper.
        for k in self.temp_wbwrapper.ws_dict.copy().keys():
            self.temp_wbwrapper.ws_dict.pop(k)
        wswr = self.temp_wbwrapper.stage_ws(
            ws_name=self.sheet_name,
            header_row=1,
            **kwargs)
        return wswr

    def clean_up(self):
        """
        (Setup method.)
        Close the WorkbookWrapper. Delete the temporary files and
        directory.
        :return: None
        """
        if self.temp_wbwrapper is not None:
            self.temp_wbwrapper.close_wb(save=False)
        self.temp_wbwrapper = None
        if self.temp_dir.exists():
            for fn in os.listdir(self.temp_dir):
                fp = self.temp_dir / fn
                fp.unlink()
            self.temp_dir.rmdir()


class UnitTest(unittest.TestCase):

    # Helper object for generating new test files and wiping old ones.
    FH = FileHandler()

    master = FH.master
    temp_dir = FH.temp_dir
    sheet_name = FH.sheet_name
    temp_fn = FH.temp_fn
    temp_fp = FH.temp_fp

    def new_copy(self):
        return self.FH.new_copy()

    def reload_wswrapper(self, **kwargs):
        return self.FH.reload_wswrapper(**kwargs)

    def clean_up(self):
        return self.FH.clean_up()

    def test_new_wbwrapper(self):
        """
        Creation of new WorkbookWrapper (and copying of master).
        """
        wb = self.new_copy()
        self.assertTrue(self.temp_fp.exists())
        self.assertTrue(wb.is_loaded)

    def test_new_wswrapper(self):
        """
        Creation of new WorksheetWrapper.
        """
        self.new_copy()
        wswr = self.reload_wswrapper()

    def test_rename(self):
        """
        Test the various methods of renaming a worksheet, and test that
        the ``.ws_dict`` in the WorkbookWrapper object has changed the
        key accordingly.
        :return:
        """
        test_name = 'temp'
        wbwp = self.new_copy()
        # Test with wbwp.rename_ws()
        wswp = wbwp.stage_ws(self.sheet_name)
        wbwp.rename_ws(self.sheet_name, test_name)
        wswp = wbwp[test_name]
        self.assertTrue(wswp.ws.title == test_name)
        self.assertFalse(wswp.ws.title == self.sheet_name)
        # Test subscripting old and new sheetnames.
        wswp = wbwp[test_name]
        with self.assertRaises(KeyError):
            wbwp[self.sheet_name]

        # Test at stage_ws().
        self.new_copy()
        wswp = self.reload_wswrapper(rename_ws=test_name)
        self.assertTrue(wswp.ws.title == test_name)
        self.assertFalse(wswp.ws.title == self.sheet_name)
        # Test subscripting old and new sheetnames.
        wswp = wbwp[test_name]
        with self.assertRaises(KeyError):
            wbwp[self.sheet_name]

        # Test after init.
        self.new_copy()
        wswp = self.reload_wswrapper()
        wswp.rename_ws(test_name)
        self.assertTrue(wswp.ws.title == test_name)
        self.assertFalse(wswp.ws.title == self.sheet_name)
        # Test subscripting old and new sheetnames.
        wswp = wbwp[test_name]
        with self.assertRaises(KeyError):
            wbwp[self.sheet_name]

    def test_is_loaded(self):
        # Test while open (assert True).
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()
        self.assertTrue(wbwp.is_loaded)
        self.assertTrue(wswp.is_loaded)
        # Test while closed (assert False).
        wbwp.close_wb(save=False)
        self.assertFalse(wbwp.is_loaded)
        self.assertFalse(wswp.is_loaded)

    # WBWP methods
    def test_delete_ws(self):
        # Test without staging.
        wbwp = self.new_copy()
        wbwp.delete_ws(self.sheet_name)
        self.assertTrue(len(wbwp.wb.sheetnames) == 0)

        # Test after staging.
        wbwp = self.new_copy()
        self.reload_wswrapper()
        wbwp.delete_ws(self.sheet_name)
        self.assertTrue(len(wbwp.wb.sheetnames) == 0)
        with self.assertRaises(KeyError):
            wbwp[self.sheet_name]

    def test_inform_subordinates(self):
        """
        Test WorkbookWrapper._inform_subordinates()
        :return:
        """
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()
        # Manually set worksheet to None to specifically test
        # _inform_subordinates().
        wswp.ws = None
        wbwp._inform_subordinates()
        self.assertTrue(isinstance(wswp.ws, Worksheet))
        # Manually set worksheet to an actual value.
        ws_holder = wswp.ws
        wbwp.close_wb(save=False)
        self.assertTrue(wswp.ws is None)
        wswp.ws = ws_holder
        # This should reset all worksheets to None.
        wbwp._inform_subordinates()
        self.assertTrue(wswp.ws is None)

    def test_mandate_loaded(self):
        wbwp = self.new_copy()
        wbwp.close_wb()
        with self.assertRaises(RuntimeError):
            wbwp.mandate_loaded()

    # WSWP methods
    def test_populate_protected_rows(self):
        """
        Test protected_rows parameter.
        :return:
        """

        def compare_test_vals(vals, ws):
            # Compare the remaining row vals against the originally
            # collected `vals`.
            remaining_vals = [
                ws[f"A{row_num}"].value for row_num in range(2, ws.max_row + 1)
            ]
            self.assertTrue(vals == remaining_vals)
            return None

        def confirm_no_fomulas(ws):
            # Confirm that no formulas were added to any protected rows.
            for row_num in range(2, ws.max_row + 1):
                self.assertTrue(ws[f"F{row_num}"].value is None)

        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()

        # Collect the values in protected rows to ensure they stay the
        # same after culling.
        ws = wswp.ws
        protected_rows = [2, 4, 6]
        test_vals = []
        for row_num in protected_rows:
            val = ws.cell(row=row_num, column=1).value
            test_vals.append(val)

        # Delete every unprotected row.
        delete_all = {'a': lambda _: True}

        # Protect rows at `.cull()`.
        wswp.cull(delete_conditions=delete_all, protected_rows=protected_rows)
        compare_test_vals(test_vals, wswp.ws)

        # Protect rows at init.
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper(protected_rows=protected_rows)
        # Delete all unprotected rows.
        wswp.cull(delete_conditions=delete_all)
        compare_test_vals(test_vals, wswp.ws)

        # Check formulas.
        wswp.add_formulas(formulas={'F': lambda _: '=1+2'})
        confirm_no_fomulas(wswp.ws)

    def test_cull(self):
        """Test .cull() method."""
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()
        wswp.cull(delete_conditions={'a': lambda x: x < 10})
        for row_num in range(2, wswp.ws.max_row + 1):
            self.assertTrue(wswp.ws.cell(row=row_num, column=1).value >= 10)

    def test_add_formulas(self):
        """Test .add_formulas method."""
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()
        wswp.add_formulas(formulas={"F": lambda row_num: f"=C{row_num}+D{row_num}"})
        self.assertTrue(wswp.ws['F3'].value == '=C3+D3')
        self.assertTrue(wswp.ws['F5'].value == '=C5+D5')

    def test_find_match_col(self):
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper()
        col_num = wswp.find_match_col(header_row=1, match_col_name='c')
        self.assertTrue(col_num == 3)
        with self.assertRaises(RuntimeError):
            wswp.find_match_col(header_row=1, match_col_name="Nope!")

    def test_modifiable_rows(self):
        wbwp = self.new_copy()
        wswp = self.reload_wswrapper(protected_rows=[2, 4, 6])
        self.assertTrue(wswp.modifiable_rows() == [3, 5])

        wswp = self.reload_wswrapper()
        self.assertTrue(wswp.modifiable_rows() == [2, 3, 4, 5, 6])

        wswp = self.reload_wswrapper()
        self.assertTrue(wswp.modifiable_rows(protected_rows=[2, 4, 6]) == [3, 5])

    def test_apply_bool_operator(self):
        both_sets = [{1, 2, 3}, {3, 4, 5}]
        wswp_class = xlsx_copycull.WorksheetWrapper
        self.assertTrue(
            wswp_class._apply_bool_operator(both_sets, 'AND') == {3})
        self.assertTrue(
            wswp_class._apply_bool_operator(both_sets, 'OR') == {1, 2, 3, 4, 5})
        self.assertTrue(
            wswp_class._apply_bool_operator(both_sets, 'XOR') == {1, 2, 4, 5})

    # Misc. functions
    def test_find_ranges(self):
        nums = [-3, -2, -1, 0, 1, 2, 3, 5, 6, 7, 8, 9, 18, 19, 20, 22]
        random.shuffle(nums)
        rges = find_ranges(nums)
        self.assertTrue(rges == [(-3, 3), (5, 9), (18, 20), (22, 22)])



if __name__ == '__main__':
    fh = FileHandler()
    # Clean up a prior failed test, if it exists.
    fh.clean_up()
    unittest.main()
    fh.clean_up()
