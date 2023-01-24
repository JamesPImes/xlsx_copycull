# Copyright (c) 2021-2023, James P. Imes. All rights reserved.

"""
Tools for generating copies of a spreadsheet and culling rows based on
some condition (e.g., delete all rows whose cell in a certain column has
a value less than 100). Can optionally add dynamic Microsoft Excel
formulas afterward.
"""

import os
import shutil
from pathlib import Path
import openpyxl


__all__ = [
    'add_formulas_to_column',
    'WorkbookWrapper',
    'WorksheetWrapper',
]


class WorkbookWrapper:
    """
    A wrapper class for openpyxl workbooks, with added methods for
    generating modified copies (e.g., reducing to only the rows relevant
    to a portion of the data).

    By design, this will leave the original spreadsheet alone and will
    generate a copy to modify.

    In particular, look into the ``.cull()`` and ``.add_formulas()``
    methods of the subordinate ``WorksheetWrapper`` objects (which
    get stored in the ``.ws_dict`` attribute).

    Before modifying any worksheet in the wrapped workbook with the
    added methods, you MUST stage it with the ``.stage_ws()`` method,
    specifying its name, the row containing the header (defaults to 1),
    and various optional parameters, such as which rows to leave alone
    (this will create a ``WorksheetWrapper`` object).

    Access the staged ``WorksheetWrapper`` objects either directly in
    the ``.ws_dict`` attribute (a dict, keyed by sheet name), or by
    subscripting on the ``WorkbookWrapper`` object (passing the sheet
    name)::

        ``some_wb_wrapper.ws_dict['Sheet1'].cull(<...>)``

            ...is equivalent to...

        ``some_wb_wrapper['Sheet1'].cull(<...>)``

    (Remember, though, that worksheets must first be staged with
    ``.stage_ws()``, or this would raise a ``KeyError``.)

    .. warning::
      As with any script that uses openpyxl to modify spreadsheets, any
      formulas that exist in the original spreadsheet will most likely
      NOT survive the insertion or deletion of rows or columns (or
      changing of worksheet names, etc.). Thus, it is highly recommended
      that you flatten all possible formulas, and use the
      ``.add_formulas()`` method in the ``WorksheetWrapper`` class to
      the extent possible for your use case.
    """

    def __init__(
            self,
            orig_fp: Path,
            copy_fp: Path = None,
            uid=None,
            no_copy=False):
        """
        A wrapper for an openpyxl Workbook object. Access the Workbook
        object directly in the ``.wb`` attribute.  The Workbook will
        be loaded at init.

        .. note::

          ``.wb`` will be set to ``None`` if the file is not currently
          open. Open it with the ``.load_wb()`` method, close it with
          ``.close_wb()`` (which will NOT save by default), and check
          whether it is currently open with the ``.is_loaded``
          property.

        :param orig_fp: Filepath to the workbook to load (and copy from).
         Must be in the ``.xlsx`` or ``.xlsm`` formats!
        :param copy_fp: Filepath at which to save the copied workbook.
         The filename should end in ``'.xlsx'`` or ``'.xlsm'``.
        :param uid: (Optional) An internal unique identifier.
        :param no_copy: Use this to modify the original spreadsheet,
         without copying. By default, the original spreadsheet will be
         copied to the filepath at ``copy_fp`` (i.e. ``no_copy=False``).

          .. warning::
            Using ``no_copy`` will irrevocably modify the original
            spreadsheet.
        """
        # a dict of subordinate WorksheetWrapper objects
        self.ws_dict = {}
        # The openpyxl workbook -- will be set to None whenever the wb
        # is NOT currently open.  (Check if this is currently set with
        # the property `self.is_loaded`)
        self.wb = None

        self.uid = uid
        self.orig_fp = Path(orig_fp)
        if no_copy:
            self.copy_fp = self.orig_fp
        elif copy_fp is None:
            raise ValueError(
                "specify `copy_fp=<path>` to create a copy, "
                "or use `no_copy=True` to modify the original file.")
        elif self.orig_fp == Path(copy_fp):
            raise ValueError(
                "Cannot copy source to its same filepath."
                "Use `no_copy=True` to modify the original file.")
        else:
            self.copy_fp = Path(copy_fp)
            self.copy_original()

    @property
    def is_loaded(self):
        return self.wb is not None

    def __getitem__(self, item):
        # Subscripting passes keys through to `.ws_dict` dict attribute.
        try:
            return self.ws_dict[item]
        except KeyError:
            raise KeyError(
                f"worksheet {item!r} has not yet been staged (or "
                f"does not exist in this workbook). "
                f"Must first call `.stage_ws()`")

    def stage_ws(
            self,
            ws_name,
            header_row: int = 1,
            first_modifiable_row: int = -1,
            protected_rows: set = None,
            rename_ws: str = None):
        """
        Prepare a worksheet for modification.

        :param ws_name: The (original) sheet name.
        :param header_row: The row containing headers (an int, indexed
         to 1)
        :param first_modifiable_row: (Optional) The first row that may
         be modified (an int, indexed to 1). If not set, will default to
         the first row after the ``header_row``.
        :param protected_rows: A list-like object containing the rows
         that should never be deleted. Row numbers before
         ``first_modifiable_row`` and the header row will be
         automatically added to the rows that may not be deleted.
        :param rename_ws: (Optional) A string, for how to rename the
         worksheet. Defaults to ``None``, in which case, it will not be
         renamed.

          .. warning::
            If the worksheet is renamed, the new name will be the key
            for this worksheet, and NOT the original worksheet name.

        :return: The ``WorksheetWrapper`` object for the newly staged
         sheet (which is also stored to ``.ws_dict``, keyed by the sheet
         name).
        """
        wswp = WorksheetWrapper(
            wb_wrapper=self,
            ws_name=ws_name,
            header_row=header_row,
            protected_rows=protected_rows,
            first_modifiable_row=first_modifiable_row)
        self.ws_dict[ws_name] = wswp
        if rename_ws:
            self.rename_ws(ws_name, new_name=rename_ws)
        return wswp

    def copy_original(self, fp=None, stage_new_fp=False) -> None:
        """
        Copy the source spreadsheet to the new filepath at ``fp``, and
        store that new filepath to ``.copy_fp``. (If ``fp`` is not
        specified here, will default to whatever is already set in
        ``.copy_fp``.)
        :param fp: The filepath to copy to.
        :param stage_new_fp: A bool, whether to set the filepath of the
         newly copied workbook as the target workbook of this
         ``WorkbookWrapper`` object. That is, whether the newly copied
         spreadsheet is the one we want to be working on. Defaults to
         ``False``.

          .. note::
            If the workbook is currently open and ``stage_new_fp=True``
            is passed, it will raise a ``RuntimeError``. To avoid that
            error, save and close the workbook first:

            .. code-block::

              workbook_wrapper.save_wb()
              workbook_wrapper.close_wb()

        :return: None
        """
        if self.is_loaded and stage_new_fp:
            raise RuntimeError(
                "Workbook is currently open. Save and close with "
                "`.close_wb()` before copying.")
        if fp is None:
            fp = self.copy_fp
        fp = Path(fp)
        os.makedirs(fp.parent, exist_ok=True)
        shutil.copy(self.orig_fp, fp)
        self.copy_fp = fp
        return None

    def delete_ws(self, ws_name):
        """
        Delete a worksheet from the workbook. (The worksheet need not be
        staged.)
        :param ws_name: The name of the worksheet to discard.
        :return: None
        """
        self.mandate_loaded()
        self.wb.remove(self.wb[ws_name])
        if ws_name in self.ws_dict.keys():
            self.ws_dict.pop(ws_name)
        return None

    def load_wb(self, **load_workbook_kwargs):
        """
        Open the workbook at the filepath stored in ``.copy_fp`` (and
        behind the scenes, inform all subordinate worksheets that they
        are now open for modification -- by setting their ``.ws``
        attributes to the appropriate openpyxl worksheet object.)

        :param load_workbook_kwargs: (Optional) Keyword arguments to
         pass through to the ``openpyxl.load_workbook()`` method. See
         documentation on ``openpyxl.load_workbook()`` for optional
         parameters.

         .. warning::
           This functionality is not strictly supported by the
           ``xlsx_copycull`` module. You may run into unexpected
           behavior or errors.

        :return: None
        """
        if self.is_loaded:
            return
        self.wb = openpyxl.load_workbook(self.copy_fp, **load_workbook_kwargs)
        # Update all of the staged worksheets.
        self._inform_subordinates()
        return None

    def _inform_subordinates(self) -> None:
        """
        INTERNAL USE:

        Inform the subordinate ``WorksheetWrapper`` objects whether the
        workbook has been opened or closed. If opened, set their ``.ws``
        attributes to their respective openpyxl worksheet.
        :return: None
        """
        is_loaded = self.is_loaded
        for ws_name, ws_wrapper in self.ws_dict.items():
            ws = None
            if is_loaded:
                ws = self.wb[ws_name]
            ws_wrapper.ws = ws
        return None

    def close_wb(self) -> None:
        """
        Close the workbook, and inform the subordinates that they cannot
        be modified until the workbook is reopened with ``.load_wb()``.
        :return: None
        """
        if not self.is_loaded:
            return None
        self.wb.close()
        self.wb = None
        # Update all of the staged worksheets.
        self._inform_subordinates()
        return None

    def save_wb(self, fp=None) -> None:
        """
        Save the ``.xlsx`` or ``.xlsm`` file.
        :param fp: The filepath at which to save the workbook. If not
         specified here, will save to the path currently configured in
         the ``.copy_fp`` attribute.
        :return: None
        """
        self.mandate_loaded()
        if fp is None:
            fp = self.copy_fp
        self.wb.save(fp)
        return None

    def mandate_loaded(self):
        """Raise an error if the ``.wb`` is not currently loaded."""
        if not self.is_loaded:
            raise RuntimeError(
                "Workbook is not currently open. Use the `.load_wb()` method.")
        return None

    def rename_ws(self, old_name, new_name):
        """
        Rename a worksheet. (Workbook must be open, and worksheet with
        ``old_name`` must already be staged.)

        Note that renaming the worksheet will also modify the
        corresponding ``.ws_dict`` key::

            ws_wrapper1 = wb_wrapper.ws_dict['Sheet1']  # OK
            ws_wrapper1 = wb_wrapper['Sheet1']  # OK
            wb_wrapper_obj.rename_ws('Sheet1', 'Prices')
            ws_wrapper1 = wb_wrapper.ws_dict['Prices']  # new sheet name
            ws_wrapper1 = wb_wrapper['Prices']  # new sheet name
            ws_wrapper1 = wb_wrapper['Sheet1']  # raises KeyError.
        """
        self.mandate_loaded()
        self.ws_dict[old_name].rename_ws(new_name=new_name)
        return None


class WorksheetWrapper:
    """
    A wrapper class for openpyxl worksheets, with added methods for
    culling rows (based on whether cell values match a specified
    condition) and adding formulas.
    """
    def __init__(
            self,
            wb_wrapper: WorkbookWrapper,
            ws_name: str,
            header_row: int = 1,
            protected_rows: set = None,
            first_modifiable_row: int = -1):
        """
        :param wb_wrapper: The parent ``WorkbookWrapper`` object.
        :param ws_name: The name of this worksheet.
        :param header_row: The row containing headers (an int, indexed
         to 1)
        :param protected_rows: (Optional) A list-like object containing
         the rows that should never be modified or deleted. Rows before
         ``first_modifiable_row`` and the header row will be
         automatically added.

         .. note::
           ``.protected_rows`` may change behind the scenes if rows are
            deleted by ``.cull()``. If rows are inserted or deleted
            outside the functionality of this module, ``.protected_rows``
            may get corrupted.

        :param first_modifiable_row: (Optional) The first row that may
         be modified or deleted (an int, indexed to 1). If not set, will
         default to the first row after the ``header_row``.
        """
        self.wb_wrapper = wb_wrapper
        self.ws_name = ws_name
        self.ws = None
        if wb_wrapper.is_loaded:
            self.ws = wb_wrapper.wb[ws_name]
        self.header_row = header_row
        self.first_modifiable_row = first_modifiable_row
        self.protected_rows = self._populate_protected_rows(
            protected_rows, first_modifiable_row)
        self.last_protected_rows = self.protected_rows

    @property
    def is_loaded(self):
        return self.ws is not None

    def mandate_loaded(self):
        """Raise an error if the ``.wb`` is not currently loaded."""
        if not self.is_loaded:
            raise RuntimeError("Workbook is not currently open")
        return None

    def _populate_protected_rows(
            self, explicitly_protected, first_modifiable_row=None) -> set:
        """
        INTERNAL USE:

        Lock down which rows may never be deleted.
        """
        header_row = self.header_row
        if first_modifiable_row is None:
            first_modifiable_row = self.first_modifiable_row
        if first_modifiable_row <= 0:
            first_modifiable_row = header_row + 1

        protected_rows = set()
        if explicitly_protected is not None:
            protected_rows = set(explicitly_protected)

        protected_rows.update(set(range(1, first_modifiable_row)))
        protected_rows.add(header_row)  # Never delete the header.
        return protected_rows

    def rename_ws(self, new_name) -> None:
        """
        Rename this worksheet.

        :param new_name: The new name for this sheet.
        :return: None
        """
        self.mandate_loaded()
        old_name = self.ws_name
        self.ws.title = new_name
        self.ws_name = new_name
        # Update the parent WBWrapper with the new sheet name, and
        # discard the old name.
        self.wb_wrapper.ws_dict[new_name] = self.wb_wrapper.ws_dict.pop(old_name)
        return None

    def cull(self, select_conditions: dict, bool_oper='AND', protected_rows=None):
        """
        Cull the spreadsheet, based on the ``select_conditions``.  If
        more than one select condition is used (i.e. more than one key
        in ``select_conditions``), specify whether to apply ``'AND'``,
        ``'OR'``, or ``'XOR'`` boolean logic to the resulting sets by
        passing one of those as ``bool_oper`` (defaults to ``'AND'``).

        .. note::

          ``protected_rows`` is a list (or set) of integers, being
          the row numbers for those rows that should NEVER be deleted
          (indexed to 1). If rows above those numbers get deleted, the
          resulting indexes in ``protected_rows`` would not be accurate,
          so this method adjusts the protected rows to their new
          position.  The resulting row numbers are stored to attribute
          ``.last_protected_rows``.  IF AND ONLY IF ``protected_rows``
          is NOT passed as an arg here, it will be pulled from
          ``.protected_rows``, and the resulting ``protected_rows`` will
          ALSO be stored to ``.protected_rows`` (in addition to
          ``.last_protected_rows``).

        (The reason for this design choice was that rows may need to be
        protected for one operation, but not for another -- but we want
        to be able to track any changes to them. This way, they can be
        accessed in ``.protected_rows`` and/or ``.last_protected_rows``
        after calling ``.cull()`` but before calling the next method,
        which might change them.)

        :param select_conditions: A dict of column_header-to-
         select_condition pairs, to determine which rows should be
         deleted. Specifically, keyed by the header of the column to
         check under, and whose value is a function to be applied to the
         value of the cell under that column, which returns a bool. If
         the function returns ``False`` (or a ``False``-like value) when
         applied to the cell's value, that row will be marked for
         deletion.

        :param bool_oper: When using more than one select conditions
         (i.e. more than one key in the dict), use this to determine
         whether to apply OR, AND, or XOR to the resulting rows to be
         selected.  Pass one of the following:  ``'AND'``, ``'OR'``,
         ``'XOR'``. (Defaults to ``'AND'``.)

        :param protected_rows: (Optional) A list-like object containing
         the rows that should never be deleted. If not specified here,
         will pull from what is set in ``.protected_rows``.  (See
         comments above regarding ``.protected_rows`` and
         ``.last_protected_rows``.)

        :return: None
        """
        self.mandate_loaded()
        if not select_conditions:
            return None

        store_protected_rows = False
        if protected_rows is not None:
            protected_rows = self._populate_protected_rows(protected_rows)
        else:
            protected_rows = self.protected_rows
            store_protected_rows = True

        header_row = self.header_row
        ws = self.ws
        all_to_keep = []
        # Apply each select condition to the appropriate column.
        for field, keepable in select_conditions.items():
            match_col = self.find_match_col(header_row, field)
            # Keep all those rows that match our criteria or are protected.
            to_keep = []
            for row_num in range(header_row + 1, ws.max_row + 1):
                if row_num in protected_rows:
                    to_keep.append(row_num)
                cell_val = ws.cell(row=row_num, column=match_col).value
                if keepable(cell_val):
                    to_keep.append(row_num)
            all_to_keep.append(set(to_keep))

        # Apply the boolean operator to determine which rows to keep.
        final_to_keep = self._apply_bool_operator(all_to_keep, bool_oper)
        final_to_keep.update(protected_rows)
        # Delete everything else.
        to_delete = set(range(1, ws.max_row + 1)) - final_to_keep

        # Convert our raw to_delete list down to a list of 2-tuples (ranges,
        # inclusive of min/max), and delete those from bottom-up.
        rges = find_ranges(to_delete)
        rges.reverse()
        for rge in rges:
            row = rge[0]
            num_rows_to_delete = rge[1] - rge[0] + 1
            ws.delete_rows(row, num_rows_to_delete)

        # Adjust any protected row numbers upward, if any higher rows
        # were deleted
        protected_rows_after_cull = []
        for row_num in protected_rows:
            for rge in rges:
                if rge[1] < row_num:
                    row_num -= (rge[1] - rge[0] + 1)
            protected_rows_after_cull.append(row_num)

        new_protected_rows = set(protected_rows_after_cull)
        self.last_protected_rows = new_protected_rows
        if store_protected_rows:
            self.protected_rows = new_protected_rows

        return None

    def find_match_col(self, header_row, match_col_name):
        """
        Find the match column number, based on its header name.

        .. note::
            Will return the first match, so avoid duplicate header names.
        """

        self.mandate_loaded()
        ws = self.ws

        try:
            match_col = [
                c.column for c in ws[header_row] if c.value == match_col_name][0]
        except IndexError:
            raise RuntimeError(
                f"ERROR! Could not find the column name {match_col_name!r} "
                f"in header row ({header_row}).")
        return match_col

    def modifiable_rows(self, protected_rows=None) -> list:
        """
        Get a list of row numbers that currently exist in the
        spreadsheet (indexed to 1), and which are NOT in
        ``protected_rows``.

        :param protected_rows: (Optional) A collection of row numbers
         (indexed to 1) that should never be deleted. If not specified
         here, will pull from what is set in ``.protected_rows``.
        """
        self.mandate_loaded()
        if protected_rows is None:
            protected_rows = self.protected_rows
        if protected_rows is None:
            protected_rows = set()
        protected_rows = self._populate_protected_rows(protected_rows)
        row_nums = [
            j for j in range(1, self.ws.max_row + 1)
            if j not in protected_rows
        ]
        return row_nums

    def add_formulas(
            self,
            formulas,
            rows=None,
            protected_rows=None,
            number_formats: dict = None,
    ) -> dict[str: list[str]]:
        """
        Add formulas to the working spreadsheet in ``.ws``.

        :param formulas: A dict keyed by column name (e.g. ``'E'``)
         whose values are a function that generates the formula, based
         on the row number.

        For example, to generate these formulas...::

            # Column R --> =F5*AB5/$S$1  (for an example row 5)
            # Column S --> =AB5*AC5  (for an example row 5)
            # ...we would pass this dict:

            formulas = {
                "R": lambda row_num: "=F{0}*AB{0}/$S$1".format(row_num),
                "S": lambda row_num: "=AB{0}*AC{0}".format(row_num)
            }

        :param rows: The rows where formulas should be added. If
         ``rows`` is specified here, it will IGNORE ``protected_rows``
         (potentially adding formulas to all rows in ``rows``, even if
         they are also in ``protected_rows``).  However, if ``rows`` is
         NOT specified, it will insert a formula into every row EXCEPT
         ``protected_rows``.

        :param protected_rows: (Optional) A list-like object containing
         the rows that should never be deleted. If not specified here,
         will pull from what is set in ``.protected_rows``. NOTE: If
         ``.cull()`` was called before this method, then the
         ``protected_rows`` may have changed since this object was
         initialized. See comments under ``.cull()`` for a more complete
         discussion.

        :param number_formats: (Optional) A dict, keyed by column
         letter, whose values are the ``'number_format'`` to apply to
         any cells in that column to which we're adding a formula.
         Reference [openpyxl's documentation](https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html)
         for possible values and built-in options for ``number_format``.

         .. example::
            from openpyxl.styles.numbers import BUILTIN_FORMATS

            # ...

            formula_formats = {
                "R": "General",
                "S": BUILTIN_FORMATS[2]  # number format of '0.00'
            }

        :return: A dict, keyed by Column letter, whose values are a list
         of the cell names that were modified (e.g.,
         ``{'A': ['A2', 'A3']}``).
        """
        if number_formats is None:
            number_formats = {}
        self.mandate_loaded()
        if formulas is None:
            return None
        if protected_rows is not None:
            protected_rows = self._populate_protected_rows(protected_rows)
        if rows is None:
            rows = self.modifiable_rows(protected_rows=protected_rows)
        modified_cells_by_column = {}
        for column, formula in formulas.items():
            num_format = number_formats.get(column, None)
            modified_cells = add_formulas_to_column(
                ws=self.ws,
                column=column,
                rows=rows,
                formula=formula,
                number_format=num_format)
            modified_cells_by_column[column] = modified_cells
        self.last_protected_rows = protected_rows
        return modified_cells_by_column

    @staticmethod
    def _apply_bool_operator(list_of_sets: list, operator: str) -> set:
        """
        INTERNAL USE:
        Apply the specified boolean operator to a list of sets.

        :param list_of_sets: A list of sets to which to apply the bool
         operator.
        :param operator: Which boolean operator to apply -- either
         ``'AND'``, ``'OR'``, or ``'XOR'``.
        :return: A set of the resulting elements.
        """
        list_of_sets = list_of_sets.copy()
        operator = operator.upper()
        final = set()
        if list_of_sets:
            final = list_of_sets.pop()
            final = final.copy()
        if operator == 'OR':
            for new_set in list_of_sets:
                final.update(new_set)
        elif operator == 'AND':
            for new_set in list_of_sets:
                final.intersection_update(new_set)
        elif operator == 'XOR':
            for new_set in list_of_sets:
                final = final ^ new_set
        else:
            raise ValueError(
                f"`operator` must be one of ['OR', 'AND', 'XOR']. "
                f"Passed {operator!r}")
        return final


def find_ranges(nums: set) -> list:
    """
    Find ranges of consecutive integers in the set. Returns a list of
    tuples, of the first and last numbers (inclusive) in each sequence.
    :param nums: A set of integers.
    :return: A list of 2-tuples of integers, being the min and max
     of each range (inclusive).
    """
    nnums = list(nums)
    nnums.sort()
    starts = [n for n in nnums if n - 1 not in nums]
    ends = [n for n in nnums if n + 1 not in nums]
    return [*zip(starts, ends)]


def add_formulas_to_column(
        ws,
        column: str,
        rows: list,
        formula,
        number_format: str = None,
) -> list[str]:
    """
    Add formulas to every row in a column in the worksheet, based on
    each row number.

    :param ws: An openpyxl worksheet.
    :param column: The column's letter name (i.e. "D").
    :param rows: A list of integers, being the row numbers where to add
     the formula. (Indexed to 1)
    :param formula: A function that will generate a formula based on the
     current row number -- such as::

        # Generates '=F5*AB5/$S$1'   (for an example row 5).
        formula = lambda row_num: "=F{0}*AB{0}/$S$1".format(row_num)

    :param number_format: (Optional) The number format to apply to each
     cell to which a formula gets written (e.g., ``'General'``).

     .. note::

        Reference [openpyxl's documentation](https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html)
        for possible values and built-in options for ``number_format``.

     .. example::

           from openpyxl.styles.numbers import BUILTIN_FORMATS
           # ...
           number_format = "General"
           # or use one of the formats in openpyxl's builtins (this is '0.00')
           number_format = BUILTIN_FORMATS[2]

    :return: A list of all cell names that were modified (e.g.,
     ``['A2', 'A3']``.)
    """
    modified_cells = []
    for row in rows:
        cell_name = f"{column}{row}"
        ws[cell_name] = formula(row)
        if number_format is not None:
            ws[cell_name].number_format = number_format
        modified_cells.append(cell_name)
    return modified_cells
